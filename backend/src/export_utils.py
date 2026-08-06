import os
import io
import pandas as pd
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

def generate_csv_export(assignments: List[Dict[str, Any]]) -> bytes:
    df = pd.DataFrame(assignments)
    output = io.StringIO()
    df.to_csv(output, index=False)
    return output.getvalue().encode("utf-8")

def generate_excel_export(
    assignments: List[Dict[str, Any]],
    kpis: Dict[str, Any],
    allocator_name: str
) -> bytes:
    wb = Workbook()
    ws_assignments = wb.active
    ws_assignments.title = "Assignments"

    # Header style
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    if assignments:
        headers = list(assignments[0].keys())
        ws_assignments.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws_assignments.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_data in assignments:
            ws_assignments.append(list(row_data.values()))

    # Metrics sheet
    ws_metrics = wb.create_sheet(title="KPI Summary")
    ws_metrics.append(["Metric", "Value"])
    ws_metrics.cell(row=1, column=1).font = header_font
    ws_metrics.cell(row=1, column=1).fill = header_fill
    ws_metrics.cell(row=1, column=2).font = header_font
    ws_metrics.cell(row=1, column=2).fill = header_fill

    for k, v in kpis.items():
        ws_metrics.append([str(k), str(v)])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_verification_log_csv(verification: Dict[str, bool], kpis: Dict[str, Any], allocator_name: str) -> bytes:
    rule_labels = {
        "rule1_no_unsafe_allocated": "Rule 1: No unsafe/quarantined battery allocated",
        "rule2_no_duplicate_battery": "Rule 2: No battery assigned to more than one vehicle",
        "rule3_no_duplicate_vehicle": "Rule 3: No vehicle receives more than one battery",
        "rule4_min_soc_satisfied": "Rule 4: Allocated battery satisfies minimum acceptable SoC",
        "rule5_metrics_reproducible": "Rule 5: Reported metrics are recomputable from submitted assignments",
    }
    rows = [
        {"check": rule_labels.get(k, k), "result": "PASS" if v else "FAIL"}
        for k, v in verification.items() if k != "all_passed"
    ]
    rows.append({"check": "Overall", "result": "PASS" if verification.get("all_passed") else "FAIL"})
    for k, v in kpis.items():
        if k == "allocator_name":
            continue
        rows.append({"check": f"KPI: {k}", "result": str(v)})

    output = io.StringIO()
    output.write(f"Verification Log - {allocator_name}\n\n")
    pd.DataFrame(rows).to_csv(output, index=False)
    return output.getvalue().encode("utf-8")


def generate_verification_log_excel(verification: Dict[str, bool], kpis: Dict[str, Any], allocator_name: str) -> bytes:
    rule_labels = {
        "rule1_no_unsafe_allocated": "Rule 1: No unsafe/quarantined battery allocated",
        "rule2_no_duplicate_battery": "Rule 2: No battery assigned to more than one vehicle",
        "rule3_no_duplicate_vehicle": "Rule 3: No vehicle receives more than one battery",
        "rule4_min_soc_satisfied": "Rule 4: Allocated battery satisfies minimum acceptable SoC",
        "rule5_metrics_reproducible": "Rule 5: Reported metrics are recomputable from submitted assignments",
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "Verification Log"
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    ws.append([f"Verification Log — {allocator_name}"])
    ws.append([])
    ws.append(["Verification Rule", "Result"])
    for col_num in range(1, 3):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font

    for k, v in verification.items():
        if k == "all_passed":
            continue
        ws.append([rule_labels.get(k, k), "PASS" if v else "FAIL"])
    ws.append(["Overall", "PASS" if verification.get("all_passed") else "FAIL"])

    ws_kpi = wb.create_sheet(title="KPI Summary")
    ws_kpi.append(["Metric", "Value"])
    for k, v in kpis.items():
        ws_kpi.append([str(k), str(v)])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_verification_log_pdf(verification: Dict[str, bool], kpis: Dict[str, Any], allocator_name: str) -> bytes:
    rule_labels = {
        "rule1_no_unsafe_allocated": "Rule 1: No unsafe/quarantined battery allocated",
        "rule2_no_duplicate_battery": "Rule 2: No battery assigned to more than one vehicle",
        "rule3_no_duplicate_vehicle": "Rule 3: No vehicle receives more than one battery",
        "rule4_min_soc_satisfied": "Rule 4: Allocated battery satisfies minimum acceptable SoC",
        "rule5_metrics_reproducible": "Rule 5: Reported metrics are recomputable from submitted assignments",
    }
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "VerificationTitle", parent=styles["Heading1"], fontSize=18, leading=22,
        textColor=colors.HexColor("#1E293B"), spaceAfter=12
    )
    story.append(Paragraph(f"Verification Log — {allocator_name}", title_style))
    story.append(Spacer(1, 10))

    overall_pass = verification.get("all_passed", False)
    story.append(Paragraph(
        f"<b>Overall Result: {'PASS' if overall_pass else 'FAIL'}</b>",
        styles["Heading2"]
    ))
    story.append(Spacer(1, 6))

    rule_data = [["Verification Rule", "Result"]]
    for k, v in verification.items():
        if k == "all_passed":
            continue
        rule_data.append([rule_labels.get(k, k), "PASS" if v else "FAIL"])

    t_rules = Table(rule_data, colWidths=[420, 80])
    t_rules.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
    ]))
    story.append(t_rules)
    story.append(Spacer(1, 15))

    story.append(Paragraph("<b>Quantitative Metrics:</b>", styles["Heading2"]))
    kpi_data = [["Metric", "Value"]]
    for k, v in kpis.items():
        kpi_data.append([str(k).replace("_", " ").title(), str(v)])
    t_kpi = Table(kpi_data, colWidths=[250, 250])
    t_kpi.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
    ]))
    story.append(t_kpi)

    doc.build(story)
    return buffer.getvalue()


def generate_pdf_export(
    assignments: List[Dict[str, Any]],
    kpis: Dict[str, Any],
    allocator_name: str
) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#1E293B"),
        spaceAfter=12
    )

    story.append(Paragraph(f"BatteryHealth Allocation Report — {allocator_name}", title_style))
    story.append(Spacer(1, 10))

    # KPI Summary Table
    story.append(Paragraph("<b>Executive Summary KPIs:</b>", styles["Heading2"]))
    kpi_data = [["Metric", "Value"]]
    for k, v in kpis.items():
        kpi_data.append([str(k).replace("_", " ").title(), str(v)])

    t_kpi = Table(kpi_data, colWidths=[250, 250])
    t_kpi.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
    ]))
    story.append(t_kpi)
    story.append(Spacer(1, 15))

    # Top Assignments Table
    story.append(Paragraph("<b>Sample Vehicle Assignments:</b>", styles["Heading2"]))
    if assignments:
        cols = ["request_id", "vehicle_type", "priority", "battery_id", "battery_soc", "battery_soh", "battery_tier"]
        ass_data = [[c.replace("_", " ").title() for c in cols]]
        for row in assignments[:25]:  # Top 25 for clean PDF fit
            ass_data.append([str(row.get(c, "")) for c in cols])

        t_ass = Table(ass_data, colWidths=[65, 85, 55, 65, 65, 65, 65])
        t_ass.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#334155')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ]))
        story.append(t_ass)

    doc.build(story)
    return buffer.getvalue()
